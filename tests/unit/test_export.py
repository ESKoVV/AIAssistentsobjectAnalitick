from __future__ import annotations

import asyncio
import csv
from datetime import UTC, datetime, timedelta
from io import BytesIO, StringIO

from openpyxl import load_workbook
from starlette.requests import Request

from apps.api.public.config import APIConfig, AuthConfig
from apps.api.public.errors import SemanticValidationError
from apps.api.public.export import CSV_HEADER, export_csv, export_xlsx
from apps.api.public.repository import SnapshotItemRecord, SnapshotRecord
import apps.api.public.server as public_server
from apps.api.public.server import create_app
from apps.api.schemas.top import (
    SamplePost,
    ScoreBreakdown,
    SourceSummary,
    TopExportQueryParams,
    TopItem,
    UrgencyLevel,
)


def build_top_item(
    *,
    rank: int = 1,
    cluster_id: str = "cluster-1",
    summary: str = "Проблема в районе",
    category: str = "housing",
    category_label: str = "ЖКХ",
    geo_regions: list[str] | None = None,
    sources: list[SourceSummary] | None = None,
) -> TopItem:
    computed_at = datetime(2026, 4, 5, 12, 0, tzinfo=UTC)
    period_start = computed_at - timedelta(hours=24)
    return TopItem(
        rank=rank,
        cluster_id=cluster_id,
        summary=summary,
        category=category,
        category_label=category_label,
        key_phrases=["фраза"],
        urgency=UrgencyLevel.HIGH,
        urgency_reason="высокий суммарный балл 0.82",
        mention_count=25,
        unique_authors=18,
        reach_total=4200,
        growth_rate=1.7,
        is_new=False,
        is_growing=True,
        geo_regions=geo_regions or ["Волгоград"],
        sources=sources
        or [
            SourceSummary(source_type="vk", count=3),
            SourceSummary(source_type="telegram", count=2),
        ],
        sample_posts=[
            SamplePost(
                doc_id=f"doc-{rank}",
                text_preview="Текст сообщения",
                source_type="vk",
                created_at=computed_at - timedelta(hours=2),
                reach=250,
                source_url="https://example.test/post/1",
            ),
        ],
        score=0.82,
        score_breakdown=ScoreBreakdown(
            volume=0.8,
            dynamics=0.7,
            sentiment=0.6,
            reach=0.5,
            geo=0.4,
            source=0.3,
        ),
        period_start=period_start,
        period_end=computed_at,
        computed_at=computed_at,
    )


def build_snapshot_item(
    *,
    rank: int = 1,
    cluster_id: str = "cluster-1",
    summary: str = "Проблема в районе",
    category: str = "housing",
    category_label: str = "ЖКХ",
) -> SnapshotItemRecord:
    return SnapshotItemRecord(
        cluster_id=cluster_id,
        rank=rank,
        score=0.82,
        summary=summary,
        category=category,
        category_label=category_label,
        key_phrases=["фраза"],
        mention_count=25,
        unique_authors=18,
        unique_sources=2,
        reach_total=4200,
        growth_rate=1.7,
        geo_regions=["Волгоград", "Волжский"],
        score_breakdown={
            "volume": 0.8,
            "dynamics": 0.7,
            "sentiment": 0.6,
            "reach": 0.5,
            "geo": 0.4,
            "source": 0.3,
        },
        sample_doc_ids=["doc-1"],
        sentiment_score=-0.3,
        is_new=False,
        is_growing=True,
        sources=[
            SourceSummary(source_type="vk", count=3),
            SourceSummary(source_type="telegram", count=2),
        ],
        sample_posts=[],
        timeline=[],
    )


def build_export_app(monkeypatch, *, items: list[SnapshotItemRecord]):
    computed_at = datetime(2026, 4, 5, 12, 0, tzinfo=UTC)

    class StubRepository:
        def __init__(self, dsn: str, *, documents_table: str = "normalized_messages") -> None:
            del dsn, documents_table

        def fetch_snapshot(self, *, period_hours: int, as_of=None):
            del period_hours, as_of
            return SnapshotRecord(
                ranking_id="ranking-1",
                computed_at=computed_at,
                period_start=computed_at - timedelta(hours=24),
                period_end=computed_at,
                top_n=10,
                period_hours=24,
            )

        def fetch_snapshot_items(self, *, ranking_id: str, cluster_ids=None, category=None):
            del ranking_id, cluster_ids
            if category is None:
                return items
            return [item for item in items if item.category == category]

    monkeypatch.setattr(public_server, "PublicAPIRepository", StubRepository)
    return create_app(
        APIConfig(
            database_url="postgresql://test:test@localhost:5432/test",
            auth=AuthConfig(disabled=True),
        ),
    )


def invoke_export_route(app, params: TopExportQueryParams):
    route = next(route for route in app.routes if getattr(route, "path", None) == "/api/v1/top/export")
    request = Request(
        {
            "type": "http",
            "method": "GET",
            "path": "/api/v1/top/export",
            "headers": [],
            "query_string": b"",
            "client": ("testclient", 123),
            "server": ("testserver", 80),
            "scheme": "http",
            "app": app,
            "root_path": "",
            "http_version": "1.1",
        },
    )
    return asyncio.run(route.endpoint(request=request, params=params, _=object()))


def test_export_csv_includes_header_and_rows() -> None:
    items = [build_top_item(rank=index, cluster_id=f"cluster-{index}") for index in range(1, 6)]

    payload = export_csv(items)
    rows = list(csv.reader(StringIO(payload.decode("utf-8"))))

    assert len(rows) == 6
    assert rows[0] == CSV_HEADER


def test_export_csv_returns_header_only_for_empty_input() -> None:
    payload = export_csv([])
    rows = list(csv.reader(StringIO(payload.decode("utf-8"))))

    assert rows == [CSV_HEADER]


def test_export_csv_joins_multiple_geo_regions_with_semicolon() -> None:
    payload = export_csv(
        [
            build_top_item(
                geo_regions=["Волгоград", "Волжский", "Камышин"],
                sources=[SourceSummary(source_type="vk", count=4)],
            ),
        ],
    )
    rows = list(csv.reader(StringIO(payload.decode("utf-8"))))

    assert rows[1][8] == "Волгоград;Волжский;Камышин"


def test_export_xlsx_creates_three_sheets() -> None:
    items = [build_top_item(rank=index, cluster_id=f"cluster-{index}") for index in range(1, 4)]

    payload = export_xlsx(items)
    workbook = load_workbook(BytesIO(payload))

    assert workbook.sheetnames == ["Топ-10", "Score breakdown", "Источники"]
    assert workbook["Топ-10"].max_row == len(items) + 1


def test_top_export_csv_endpoint_returns_attachment(monkeypatch) -> None:
    app = build_export_app(
        monkeypatch,
        items=[
            build_snapshot_item(
                rank=1,
                cluster_id="cluster-housing",
                summary="Проблемы с вывозом мусора",
                category="housing",
                category_label="ЖКХ",
            ),
            build_snapshot_item(
                rank=2,
                cluster_id="cluster-roads",
                summary="Яма на дороге",
                category="roads",
                category_label="Дороги и транспорт",
            ),
        ],
    )

    response = invoke_export_route(app, TopExportQueryParams(period="24h", format="csv", category="housing"))

    assert response.status_code == 200
    assert response.headers["content-type"].startswith("text/csv")
    assert response.headers["content-disposition"] == "attachment; filename=top10_2026-04-05.csv"
    rows = list(csv.reader(StringIO(response.body.decode("utf-8"))))
    assert len(rows) == 2
    assert rows[1][1] == "cluster-housing"
    assert rows[1][3] == "housing"


def test_top_export_xlsx_endpoint_returns_attachment(monkeypatch) -> None:
    app = build_export_app(monkeypatch, items=[build_snapshot_item()])
    response = invoke_export_route(app, TopExportQueryParams(period="24h", format="xlsx"))

    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    assert response.headers["content-disposition"] == "attachment; filename=top10_2026-04-05.xlsx"


def test_top_export_rejects_unsupported_format(monkeypatch) -> None:
    app = build_export_app(monkeypatch, items=[build_snapshot_item()])

    try:
        invoke_export_route(app, TopExportQueryParams(period="24h", format="pdf"))
    except SemanticValidationError as exc:
        assert exc.status_code == 422
    else:
        raise AssertionError("SemanticValidationError was not raised for unsupported export format")
