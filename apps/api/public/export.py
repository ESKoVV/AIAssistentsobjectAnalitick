from __future__ import annotations

import csv
from io import BytesIO, StringIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from apps.api.schemas.top import TopItem


CSV_HEADER = [
    "rank",
    "cluster_id",
    "summary",
    "category",
    "mention_count",
    "unique_authors",
    "reach_total",
    "growth_rate",
    "geo_regions",
    "sources",
    "score",
    "urgency",
    "period_start",
    "period_end",
    "computed_at",
]

BREAKDOWN_HEADER = ["rank", "cluster_id", "volume", "dynamics", "sentiment", "reach", "geo", "source", "итого"]
SOURCES_HEADER = ["rank", "cluster_id", "source_type", "count"]

URGENCY_FILLS = {
    "critical": PatternFill(fill_type="solid", fgColor="FECACA"),
    "high": PatternFill(fill_type="solid", fgColor="FED7AA"),
    "medium": PatternFill(fill_type="solid", fgColor="FEF08A"),
    "low": PatternFill(fill_type="solid", fgColor="E5E7EB"),
}


def export_csv(items: list[TopItem]) -> bytes:
    buffer = StringIO()
    writer = csv.writer(buffer)
    writer.writerow(CSV_HEADER)
    for item in items:
        writer.writerow(_top_row(item))
    return buffer.getvalue().encode("utf-8")


def export_xlsx(items: list[TopItem]) -> bytes:
    workbook = Workbook()
    top_sheet = workbook.active
    top_sheet.title = "Топ-10"
    _append_sheet(top_sheet, CSV_HEADER, [_top_row(item) for item in items])
    _style_header(top_sheet)
    _style_urgency_column(top_sheet, len(items))

    breakdown_sheet = workbook.create_sheet("Score breakdown")
    _append_sheet(
        breakdown_sheet,
        BREAKDOWN_HEADER,
        [
            [
                item.rank,
                item.cluster_id,
                item.score_breakdown.volume,
                item.score_breakdown.dynamics,
                item.score_breakdown.sentiment,
                item.score_breakdown.reach,
                item.score_breakdown.geo,
                item.score_breakdown.source,
                item.score,
            ]
            for item in items
        ],
    )
    _style_header(breakdown_sheet)

    sources_sheet = workbook.create_sheet("Источники")
    source_rows = [
        [item.rank, item.cluster_id, source.source_type, source.count]
        for item in items
        for source in item.sources
    ]
    _append_sheet(sources_sheet, SOURCES_HEADER, source_rows)
    _style_header(sources_sheet)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _append_sheet(sheet, header: list[str], rows: list[list[object]]) -> None:  # type: ignore[no-untyped-def]
    sheet.append(header)
    for row in rows:
        sheet.append(row)


def _style_header(sheet) -> None:  # type: ignore[no-untyped-def]
    for cell in sheet[1]:
        cell.font = Font(bold=True)


def _style_urgency_column(sheet, item_count: int) -> None:  # type: ignore[no-untyped-def]
    urgency_col = 12
    for row_index in range(2, item_count + 2):
        cell = sheet.cell(row=row_index, column=urgency_col)
        fill = URGENCY_FILLS.get(str(cell.value))
        if fill is not None:
            cell.fill = fill


def _top_row(item: TopItem) -> list[object]:
    return [
        item.rank,
        item.cluster_id,
        item.summary,
        item.category,
        item.mention_count,
        item.unique_authors,
        item.reach_total,
        item.growth_rate,
        ";".join(item.geo_regions),
        ";".join(f"{source.source_type}:{source.count}" for source in item.sources),
        item.score,
        item.urgency.value,
        item.period_start.isoformat(),
        item.period_end.isoformat(),
        item.computed_at.isoformat(),
    ]
